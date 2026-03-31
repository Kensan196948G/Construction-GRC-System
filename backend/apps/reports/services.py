"""適用宣言書（SoA）生成サービス

ISO27001:2022 全93管理策の適用宣言書（Statement of Applicability）を
Excel/PDFデータとして生成するサービスレイヤー。
"""

from __future__ import annotations

import io
from datetime import UTC, datetime
from typing import Any

from django.utils import timezone

from apps.controls.models import ISO27001Control

# ISO27001:2022 全93管理策のドメイン別管理策数
ISO27001_DOMAIN_CONTROL_COUNTS: dict[str, int] = {
    "organizational": 37,
    "people": 8,
    "physical": 14,
    "technological": 34,
}

# ドメイン別管理策IDレンジ（A.5〜A.8）
ISO27001_DOMAIN_PREFIXES: dict[str, str] = {
    "organizational": "A.5",
    "people": "A.6",
    "physical": "A.7",
    "technological": "A.8",
}


class SoAGenerator:
    """適用宣言書（Statement of Applicability）生成サービス"""

    @classmethod
    def _build_soa_data(cls) -> list[dict[str, Any]]:
        """全管理策のSoAデータを構築する。

        DBに登録されている管理策を取得し、SoA用のデータ構造に変換する。

        Returns:
            管理策ごとのSoAレコードリスト
        """
        controls = ISO27001Control.objects.all().order_by("control_id")
        soa_rows: list[dict[str, Any]] = []

        for ctrl in controls:
            domain_label = ctrl.get_domain_display() if hasattr(ctrl, "get_domain_display") else ctrl.domain
            status_label = (
                ctrl.get_implementation_status_display()
                if hasattr(ctrl, "get_implementation_status_display")
                else ctrl.implementation_status
            )

            soa_rows.append({
                "control_id": ctrl.control_id,
                "domain": ctrl.domain,
                "domain_display": domain_label,
                "title": ctrl.title,
                "description": ctrl.description,
                "is_applicable": ctrl.is_applicable,
                "exclusion_reason": ctrl.exclusion_reason if not ctrl.is_applicable else "",
                "implementation_status": ctrl.implementation_status,
                "implementation_status_display": status_label,
                "implementation_notes": ctrl.implementation_notes,
                "owner": str(ctrl.owner) if ctrl.owner else "",
                "evidence_required": ctrl.evidence_required,
                "nist_csf_mapping": ctrl.nist_csf_mapping,
                "last_reviewed_at": ctrl.last_reviewed_at,
            })

        return soa_rows

    @classmethod
    def _get_summary_statistics(
        cls, soa_data: list[dict[str, Any]]
    ) -> dict[str, Any]:
        """SoAデータからサマリー統計を算出する。

        Returns:
            ドメイン別・ステータス別の集計結果
        """
        total = len(soa_data)
        applicable = sum(1 for r in soa_data if r["is_applicable"])
        not_applicable = total - applicable

        status_counts: dict[str, int] = {}
        domain_counts: dict[str, dict[str, int]] = {}

        for row in soa_data:
            status = row["implementation_status"]
            status_counts[status] = status_counts.get(status, 0) + 1

            domain = row["domain"]
            if domain not in domain_counts:
                domain_counts[domain] = {"total": 0, "implemented": 0, "applicable": 0}
            domain_counts[domain]["total"] += 1
            if row["is_applicable"]:
                domain_counts[domain]["applicable"] += 1
            if row["implementation_status"] == "implemented":
                domain_counts[domain]["implemented"] += 1

        return {
            "total_controls": total,
            "applicable": applicable,
            "not_applicable": not_applicable,
            "by_status": status_counts,
            "by_domain": domain_counts,
            "generated_at": timezone.now().isoformat(),
        }

    @classmethod
    def generate_excel(cls) -> io.BytesIO:
        """SoA Excel (xlsx) ファイルを生成する。

        Returns:
            Excel ファイルの BytesIO ストリーム
        """
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        soa_data = cls._build_soa_data()
        summary = cls._get_summary_statistics(soa_data)

        wb = Workbook()

        # --- Sheet 1: SoA 本体 ---
        ws = wb.active
        ws.title = "適用宣言書（SoA）"

        # ヘッダースタイル
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # タイトル行
        ws.merge_cells("A1:H1")
        title_cell = ws["A1"]
        title_cell.value = f"ISO27001:2022 適用宣言書（SoA） - 生成日: {datetime.now(tz=UTC).strftime('%Y-%m-%d')}"
        title_cell.font = Font(bold=True, size=14)

        # ヘッダー
        headers = [
            "管理策ID", "ドメイン", "管理策名", "適用",
            "除外理由", "実施状況", "備考", "担当者",
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # データ行
        applicable_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        not_applicable_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

        for row_idx, row_data in enumerate(soa_data, start=4):
            fill = applicable_fill if row_data["is_applicable"] else not_applicable_fill
            values = [
                row_data["control_id"],
                row_data["domain_display"],
                row_data["title"],
                "適用" if row_data["is_applicable"] else "除外",
                row_data["exclusion_reason"],
                row_data["implementation_status_display"],
                row_data["implementation_notes"],
                row_data["owner"],
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True)

        # 列幅調整
        column_widths = [12, 18, 40, 8, 25, 18, 30, 15]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + col_idx)].width = width

        # --- Sheet 2: サマリー ---
        ws_summary = wb.create_sheet("サマリー")
        ws_summary["A1"] = "SoA サマリー統計"
        ws_summary["A1"].font = Font(bold=True, size=14)

        ws_summary["A3"] = "管理策総数"
        ws_summary["B3"] = summary["total_controls"]
        ws_summary["A4"] = "適用"
        ws_summary["B4"] = summary["applicable"]
        ws_summary["A5"] = "除外"
        ws_summary["B5"] = summary["not_applicable"]

        row = 7
        ws_summary.cell(row=row, column=1, value="ドメイン").font = Font(bold=True)
        ws_summary.cell(row=row, column=2, value="合計").font = Font(bold=True)
        ws_summary.cell(row=row, column=3, value="適用").font = Font(bold=True)
        ws_summary.cell(row=row, column=4, value="実施済み").font = Font(bold=True)
        row += 1
        for domain, counts in summary["by_domain"].items():
            ws_summary.cell(row=row, column=1, value=domain)
            ws_summary.cell(row=row, column=2, value=counts["total"])
            ws_summary.cell(row=row, column=3, value=counts["applicable"])
            ws_summary.cell(row=row, column=4, value=counts["implemented"])
            row += 1

        # 出力
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @classmethod
    def generate_pdf_data(cls) -> dict[str, Any]:
        """SoA PDF生成用のデータ構造を返す。

        WeasyPrint等で描画するためのコンテキストデータを構築する。

        Returns:
            PDF テンプレートに渡すコンテキスト辞書:
            {
                "title": str,
                "generated_at": str,
                "controls": [...],
                "summary": {...},
                "domains": [...],
            }
        """
        soa_data = cls._build_soa_data()
        summary = cls._get_summary_statistics(soa_data)

        # ドメイン別にグループ化
        domains: dict[str, list[dict[str, Any]]] = {}
        for row in soa_data:
            domain = row["domain_display"]
            if domain not in domains:
                domains[domain] = []
            domains[domain].append(row)

        domain_sections: list[dict[str, Any]] = [
            {"domain_name": name, "controls": ctrls}
            for name, ctrls in domains.items()
        ]

        return {
            "title": "ISO27001:2022 適用宣言書（Statement of Applicability）",
            "generated_at": timezone.now().strftime("%Y年%m月%d日 %H:%M"),
            "controls": soa_data,
            "summary": summary,
            "domains": domain_sections,
        }
